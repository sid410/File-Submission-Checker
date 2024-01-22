import os
from openpyxl import Workbook, load_workbook
import argparse
import datetime

parser = argparse.ArgumentParser()
parser.add_argument("class_file", type=str)
parser.add_argument("activity_folder", type=str)
# parser.add_argument("submit_day", type=str)
parser.add_argument('check_list', nargs='+', default=[])
parser.add_argument("-n", "--num_students", type=int, default=40)
args = parser.parse_args()

# Convert all string in list to uppercase
def uppercase_stringlist(str_list):
    for n, f in enumerate(str_list):
        str_list[n] = f.upper()

    return str_list

# path of the folder and files to check
folder_to_check = args.activity_folder
activity_to_check = os.path.basename(folder_to_check)

files_to_check = args.check_list
files_to_check = uppercase_stringlist(files_to_check)
files_to_check.sort()


# initialize workbook, worksheet, and the submissions dictionary
def init_checking(excel_file, activity):

    Submissions = {}
    wb = load_workbook(excel_file)
    ids = wb['Classlist']

    # Create a submissions dictionary for each student
    for row in ids.iter_rows(min_row=2, min_col=1, max_row= (args.num_students + 1), max_col=1):
        for cell in row:
            id = ids.cell(row=cell.row, column=cell.column).value
            Submissions[id] = {}

    # Create worksheet and place appropriate labels
    ws = wb.create_sheet(activity + "_" + str(datetime.date.today()))
    ws.column_dimensions['A'].width = 20
    ws.cell(1, 1, "Student Number")
    for i, file in enumerate(files_to_check):
        ws.cell(1, 2+i, file)

    return wb, ws, Submissions


# Check the submission folder of one student
# def check_submission(Submissions, id, directory, submit_day):
def check_submission(Submissions, id, directory):

    # change below so that it can directory path for both Linux and Windows
    directory += "/" + str(id)

    # Check if the student made a folder for their submission using their ID
    if os.path.exists(directory):
        files_submitted = [os.path.splitext(filename)[0] for filename in os.listdir(directory)]
        files_submitted = uppercase_stringlist(files_submitted)
    else:
        files_submitted = {}
        # return Submissions 

    # Check inside their folder if they have submitted the appropriate files
    for file in files_to_check:
        if file in files_submitted:
            Submissions[id][file] = "✓"
        else:
            Submissions[id][file] = "☒"

    return Submissions

# Save the record of checking submission
def save_submissions(Submissions, wb, excel_file):

    for r, id in enumerate(Submissions):
        ws.cell(r+2, 1, id)

        for c, file in enumerate(Submissions[id]):
            ws.cell(r+2, c+2, Submissions[id][file])

    wb.save(excel_file)


if __name__ == "__main__":
    wb, ws, sub = init_checking(args.class_file, activity_to_check)

    for student in sub:
        # check_submission(sub, student, folder_to_check, args.submit_day)
        check_submission(sub, student, folder_to_check)

    save_submissions(sub, wb, args.class_file)